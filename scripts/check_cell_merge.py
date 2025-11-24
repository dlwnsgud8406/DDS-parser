#!/usr/bin/env python3
"""원본 Excel 파일의 셀 병합 상태 확인"""

from openpyxl import load_workbook

def check_merged_cells():
    """원본 Excel의 셀 병합 확인"""
    
    print("=" * 80)
    print("원본 Excel 셀 병합 상태 확인")
    print("=" * 80)
    
    # 원본 파일 읽기
    wb = load_workbook('output/shm_all_fixed.xlsx')
    
    # 첫 번째 노드 시트
    sheet_name = [s for s in wb.sheetnames if s.startswith('Node_')][0]
    ws = wb[sheet_name]
    
    print(f"\n시트: {sheet_name}")
    print(f"병합된 셀 범위 개수: {len(ws.merged_cells.ranges)}")
    
    # 처음 10개 병합 범위 출력
    print("\n처음 10개 병합 범위:")
    for i, merged_range in enumerate(list(ws.merged_cells.ranges)[:10]):
        print(f"  {i+1}. {merged_range}")
    
    # idx와 TimeInterval 컬럼의 병합 패턴 확인
    print("\nidx와 TimeInterval 컬럼 병합 패턴:")
    
    # A열(idx)과 B열(TimeInterval)의 처음 20개 셀 값 확인
    print("\n처음 20행의 값:")
    print(f"{'행':>3} | {'A(idx)':>10} | {'B(TimeInterval)':>15}")
    print("-" * 45)
    
    for row in range(1, 21):
        cell_a = ws.cell(row=row, column=1)
        cell_b = ws.cell(row=row, column=2)
        print(f"{row:3d} | {str(cell_a.value):>10} | {str(cell_b.value):>15}")

if __name__ == '__main__':
    check_merged_cells()
