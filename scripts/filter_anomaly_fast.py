#!/usr/bin/env python3
"""
원본 Excel을 openpyxl로 직접 읽어서 필요한 컬럼만 복사 (빠른 버전)
"""

from openpyxl import load_workbook, Workbook
from pathlib import Path
import sys

# Anomaly detection에 필요한 13개 PID
REQUIRED_PIDS = [
    'PID_TOPIC_NAME',           # 0x0005
    'PID_TYPE_NAME',            # 0x0007
    'PID_RELIABILITY',          # 0x001A
    'PID_DURABILITY',           # 0x001D
    'PID_LIVELINESS',           # 0x001B
    'PID_DEADLINE',             # 0x0023
    'PID_LATENCY_BUDGET',       # 0x0027
    'PID_DESTINATION_ORDER',    # 0x0025
    'PID_USER_DATA',            # 0x002C
    'PID_HISTORY',              # 0x0040
    'PID_RESOURCE_LIMIT',       # 0x0041
    'PID_TYPE_MAX_SIZE_SERIALIZED',  # 0x0061
    'PID_STATUS_INFO',          # 0x0071
]


def get_column_indices_to_copy(header_row, required_pids):
    """
    복사할 컬럼 인덱스 계산
    
    Args:
        header_row: 헤더 행 (openpyxl row)
        required_pids: 필요한 PID 목록
    
    Returns:
        복사할 컬럼 인덱스 리스트 (1-based)
    """
    # 헤더를 리스트로 변환
    headers = [cell.value for cell in header_row]
    
    # 기본 컬럼 (idx, TimeInterval, topic)
    column_indices = []
    
    if 'idx' in headers:
        column_indices.append(headers.index('idx') + 1)
    if 'TimeInterval' in headers:
        column_indices.append(headers.index('TimeInterval') + 1)
    if 'topic' in headers:
        column_indices.append(headers.index('topic') + 1)
    
    # 각 PID와 다음 Unnamed 컬럼
    for pid in required_pids:
        if pid in headers:
            pid_idx = headers.index(pid)
            column_indices.append(pid_idx + 1)  # PID 컬럼 (1-based)
            if pid_idx + 1 < len(headers):
                column_indices.append(pid_idx + 2)  # Unnamed 컬럼 (1-based)
    
    return sorted(column_indices)


def copy_columns(source_ws, target_ws, column_indices):
    """
    원본 워크시트에서 특정 컬럼들을 대상 워크시트로 복사
    
    Args:
        source_ws: 원본 워크시트
        target_ws: 대상 워크시트
        column_indices: 복사할 컬럼 인덱스 리스트 (1-based)
    """
    # 모든 행 순회
    for row_idx, row in enumerate(source_ws.iter_rows(), start=1):
        # 선택된 컬럼만 복사
        for target_col_idx, source_col_idx in enumerate(column_indices, start=1):
            source_cell = row[source_col_idx - 1]
            target_cell = target_ws.cell(row=row_idx, column=target_col_idx)
            
            # 값 복사
            target_cell.value = source_cell.value
            
            # 스타일 복사 (첫 2행만)
            if row_idx <= 2:
                if source_cell.has_style:
                    target_cell.font = source_cell.font.copy()
                    target_cell.fill = source_cell.fill.copy()
                    target_cell.border = source_cell.border.copy()
                    target_cell.alignment = source_cell.alignment.copy()


def copy_merged_cells(source_ws, target_ws, column_mapping):
    """
    셀 병합 복사 (A, B, C 컬럼만)
    
    Args:
        source_ws: 원본 워크시트
        target_ws: 대상 워크시트
        column_mapping: 원본 컬럼 -> 대상 컬럼 매핑 딕셔너리
    """
    for merged_range in source_ws.merged_cells.ranges:
        # 병합 범위 분석
        min_col = merged_range.min_col
        max_col = merged_range.max_col
        min_row = merged_range.min_row
        max_row = merged_range.max_row
        
        # A, B, C 컬럼만 복사
        if min_col == max_col and min_col in column_mapping:
            target_col = column_mapping[min_col]
            # 대상 워크시트에 병합 적용
            target_ws.merge_cells(
                start_row=min_row,
                start_column=target_col,
                end_row=max_row,
                end_column=target_col
            )


def filter_excel_fast(input_path, output_path):
    """
    openpyxl로 직접 읽어서 빠르게 필터링
    """
    print("=" * 80)
    print("Anomaly Detection용 Excel 필터링 (고속 버전)")
    print("=" * 80)
    
    # 원본 파일 읽기
    print(f"\n원본 파일 읽기: {input_path}")
    wb_source = load_workbook(input_path, data_only=False)
    
    print(f"시트 개수: {len(wb_source.sheetnames)}")
    
    # 새 Workbook 생성
    wb_target = Workbook()
    wb_target.remove(wb_target.active)
    
    # 각 시트 처리
    for sheet_name in wb_source.sheetnames:
        if sheet_name == 'Overview':
            print(f"건너뛰기: {sheet_name}")
            continue
        
        print(f"처리 중: {sheet_name}...", end=' ', flush=True)
        
        ws_source = wb_source[sheet_name]
        ws_target = wb_target.create_sheet(title=sheet_name)
        
        # 헤더 행 읽기
        header_row = list(ws_source[1])
        
        # 복사할 컬럼 인덱스 계산
        column_indices = get_column_indices_to_copy(header_row, REQUIRED_PIDS)
        
        print(f"{len(column_indices)} 컬럼...", end=' ', flush=True)
        
        # 컬럼 복사
        copy_columns(ws_source, ws_target, column_indices)
        
        # 셀 병합 복사 (A, B, C 컬럼)
        column_mapping = {}
        for target_idx, source_idx in enumerate(column_indices[:3], start=1):
            column_mapping[source_idx] = target_idx
        
        copy_merged_cells(ws_source, ws_target, column_mapping)
        
        print("완료")
    
    # 파일 저장
    print(f"\n저장 중: {output_path}")
    wb_target.save(output_path)
    
    print("\n" + "=" * 80)
    print("완료!")
    print("=" * 80)


def main():
    """메인 함수"""
    input_path = Path('output/shm_nodes_fixed.xlsx')
    output_path = Path('output/anomaly_detection_filtered.xlsx')
    
    if not input_path.exists():
        print(f"오류: 입력 파일을 찾을 수 없습니다: {input_path}")
        sys.exit(1)
    
    filter_excel_fast(input_path, output_path)
    print(f"\n생성된 파일: {output_path}")


if __name__ == '__main__':
    main()
