#!/usr/bin/env python3
"""
원본 Excel 파일에서 anomaly detection에 필요한 13개 PID만 필터링하여 새로운 Excel 생성

원본 구조를 그대로 유지:
- 8개 노드 시트
- Timestamp/Value 쌍 구조
- idx, TimeInterval 컬럼 유지
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path
import sys

# Anomaly detection에 필요한 13개 PID
REQUIRED_PIDS = [
    'PID_TOPIC_NAME',           # 0x0005
    'PID_TYPE_NAME',            # 0x0007
    'PID_RELIABILITY',          # 0x001A
    'PID_DURABILITY',           # 0x001D
    'PID_LIVELINESS',           # 0x001B
    'PID_DEADLINE',             # 0x0023
    'PID_LATENCY_BUDGET',       # 0x0027
    'PID_DESTINATION_ORDER',    # 0x0025
    'PID_USER_DATA',            # 0x002C
    'PID_HISTORY',              # 0x0040
    'PID_RESOURCE_LIMIT',       # 0x0041
    'PID_TYPE_MAX_SIZE_SERIALIZED',  # 0x0061
    'PID_STATUS_INFO',          # 0x0071
]


def filter_columns(df, required_pids):
    """
    DataFrame에서 필요한 PID와 해당 Value 컬럼만 필터링
    
    Args:
        df: 원본 DataFrame
        required_pids: 필요한 PID 이름 리스트
    
    Returns:
        필터링된 DataFrame
    """
    columns = df.columns.tolist()
    
    # 기본 컬럼 (idx, TimeInterval)
    selected_cols = ['idx', 'TimeInterval']
    
    # 각 PID와 해당 Unnamed 컬럼 찾기
    for pid in required_pids:
        if pid in columns:
            pid_idx = columns.index(pid)
            # PID 컬럼과 다음 Unnamed 컬럼 추가
            selected_cols.append(columns[pid_idx])
            if pid_idx + 1 < len(columns):
                selected_cols.append(columns[pid_idx + 1])
    
    return df[selected_cols]


def get_merge_ranges(df, column_name):
    """
    DataFrame에서 동일한 값을 가진 연속된 행의 병합 범위 계산
    
    Args:
        df: pandas DataFrame
        column_name: 병합할 컬럼 이름
    
    Returns:
        병합 범위 리스트 [(start_row, end_row), ...]
    """
    if column_name not in df.columns:
        return []
    
    merge_ranges = []
    values = df[column_name].values
    
    if len(values) == 0:
        return []
    
    start_idx = 0
    current_value = values[0]
    
    for i in range(1, len(values)):
        # NaN 처리: pandas의 isna()와 동일한 값 비교
        current_is_nan = pd.isna(current_value)
        next_is_nan = pd.isna(values[i])
        
        # 값이 변경되었는지 확인 (NaN은 NaN끼리 같은 것으로 처리)
        values_differ = (current_is_nan != next_is_nan) or \
                       (not current_is_nan and not next_is_nan and values[i] != current_value)
        
        if values_differ:
            # 병합 범위가 2개 이상의 행인 경우만 추가
            if i - start_idx > 1:
                # +3: 1행(헤더) + 2행(서브헤더) + 0-based index
                merge_ranges.append((start_idx + 3, i + 2))
            start_idx = i
            current_value = values[i]
    
    # 마지막 범위 처리
    if len(values) - start_idx > 1:
        merge_ranges.append((start_idx + 3, len(values) + 2))
    
    return merge_ranges


def apply_cell_merges(ws, df):
    """
    워크시트에 셀 병합 적용 (DataFrame 기반으로 빠르게)
    
    Args:
        ws: openpyxl worksheet
        df: pandas DataFrame
    """
    # idx 컬럼 (A열) 병합
    idx_ranges = get_merge_ranges(df, 'idx')
    for start, end in idx_ranges:
        ws.merge_cells(f'A{start}:A{end}')
    
    # TimeInterval 컬럼 (B열) 병합
    time_ranges = get_merge_ranges(df, 'TimeInterval')
    for start, end in time_ranges:
        ws.merge_cells(f'B{start}:B{end}')


def style_worksheet(ws):
    """
    워크시트에 스타일 적용 (원본 스타일 모방)
    
    Args:
        ws: openpyxl worksheet
    """
    # 헤더 스타일
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    # 테두리 스타일
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 첫 번째 행 (헤더) 스타일 적용
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # 두 번째 행 (Timestamp/Value 서브헤더) 스타일 적용
    subheader_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    subheader_font = Font(bold=True)
    
    for cell in ws[2]:
        cell.fill = subheader_fill
        cell.font = subheader_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # 컬럼 너비 자동 조정
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)  # 최대 50으로 제한
        ws.column_dimensions[column_letter].width = adjusted_width


def filter_excel(input_path, output_path):
    """
    원본 Excel에서 필요한 PID만 필터링하여 새 Excel 생성
    
    Args:
        input_path: 원본 Excel 파일 경로
        output_path: 출력 Excel 파일 경로
    """
    print("=" * 80)
    print("Anomaly Detection용 Excel 필터링")
    print("=" * 80)
    
    # 원본 파일 읽기
    print(f"\n원본 파일 읽기: {input_path}")
    xl_file = pd.ExcelFile(input_path)
    
    print(f"시트 개수: {len(xl_file.sheet_names)}")
    print(f"시트 목록: {xl_file.sheet_names}\n")
    
    # 새 Workbook 생성
    wb = Workbook()
    wb.remove(wb.active)  # 기본 시트 제거
    
    # 각 시트 처리
    total_original_cols = 0
    total_filtered_cols = 0
    
    for sheet_name in xl_file.sheet_names:
        # Overview 시트는 건너뛰기 (구조가 다름)
        if sheet_name == 'Overview':
            print(f"건너뛰기: {sheet_name} (메타데이터 시트)\n")
            continue
            
        print(f"처리 중: {sheet_name}")
        
        # 시트 읽기
        df = pd.read_excel(input_path, sheet_name=sheet_name)
        original_cols = len(df.columns)
        
        # 컬럼 필터링
        filtered_df = filter_columns(df, REQUIRED_PIDS)
        filtered_cols = len(filtered_df.columns)
        
        print(f"  원본 컬럼: {original_cols}")
        print(f"  필터링 후: {filtered_cols}")
        print(f"  감소율: {(1 - filtered_cols/original_cols)*100:.1f}%\n")
        
        total_original_cols += original_cols
        total_filtered_cols += filtered_cols
        
        # 새 시트에 쓰기
        ws = wb.create_sheet(title=sheet_name)
        
        # DataFrame을 openpyxl worksheet에 쓰기
        for r_idx, row in enumerate(filtered_df.values, start=1):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        # 헤더 쓰기
        for c_idx, col_name in enumerate(filtered_df.columns, start=1):
            ws.cell(row=1, column=c_idx, value=col_name)
        
        # 스타일 적용
        style_worksheet(ws)
        
        # 셀 병합 적용 (DataFrame 기반으로 빠르게)
        apply_cell_merges(ws, filtered_df)
    
    # 파일 저장
    print(f"저장 중: {output_path}")
    wb.save(output_path)
    
    print("\n" + "=" * 80)
    print("완료!")
    print("=" * 80)
    print(f"\n전체 통계:")
    print(f"  평균 원본 컬럼: {total_original_cols / len(xl_file.sheet_names):.0f}")
    print(f"  평균 필터링 후: {total_filtered_cols / len(xl_file.sheet_names):.0f}")
    print(f"  평균 감소율: {(1 - total_filtered_cols/total_original_cols)*100:.1f}%")
    
    # 포함된 PID 목록 출력
    print(f"\n포함된 13개 PID:")
    for i, pid in enumerate(REQUIRED_PIDS, 1):
        print(f"  {i:2d}. {pid}")


def main():
    """메인 함수"""
    # 경로 설정
    input_path = Path('output/shm_all_fixed.xlsx')
    output_path = Path('output/anomaly_detection_filtered.xlsx')
    
    # 입력 파일 존재 확인
    if not input_path.exists():
        print(f"오류: 입력 파일을 찾을 수 없습니다: {input_path}")
        sys.exit(1)
    
    # 필터링 실행
    filter_excel(input_path, output_path)
    
    print(f"\n생성된 파일: {output_path}")


if __name__ == '__main__':
    main()
