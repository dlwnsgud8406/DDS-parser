#!/usr/bin/env python3
"""shm_nodes_fixed.xlsx와 anomaly_detection_filtered.xlsx 정밀 비교"""

import pandas as pd
from openpyxl import load_workbook

def precise_comparison():
    """두 파일의 구조를 정밀하게 비교"""
    
    print("=" * 80)
    print("정밀 구조 비교: shm_nodes_fixed.xlsx vs anomaly_detection_filtered.xlsx")
    print("=" * 80)
    
    # 동일한 시트로 비교 (Unknown)
    sheet_name = 'Unknown'
    
    # 원본 파일
    print(f"\n[원본: shm_nodes_fixed.xlsx - {sheet_name}]")
    df_original = pd.read_excel('output/shm_nodes_fixed.xlsx', sheet_name=sheet_name, nrows=10)
    print(f"총 컬럼 수: {len(df_original.columns)}")
    print(f"\n처음 10개 컬럼:")
    for i, col in enumerate(df_original.columns[:10]):
        print(f"  {i:2d}. {col}")
    
    print(f"\n처음 5행 데이터 (처음 6개 컬럼):")
    print(df_original.iloc[:5, :6])
    
    # openpyxl로 셀 병합 확인
    wb_original = load_workbook('output/shm_nodes_fixed.xlsx')
    ws_original = wb_original[sheet_name]
    print(f"\n셀 병합 개수: {len(ws_original.merged_cells.ranges)}")
    print(f"처음 5개 병합 범위:")
    for i, mr in enumerate(list(ws_original.merged_cells.ranges)[:5]):
        print(f"  {i+1}. {mr}")
    
    # 필터링된 파일
    print(f"\n[필터링: anomaly_detection_filtered.xlsx - {sheet_name}]")
    df_filtered = pd.read_excel('output/anomaly_detection_filtered.xlsx', sheet_name=sheet_name, nrows=10)
    print(f"총 컬럼 수: {len(df_filtered.columns)}")
    print(f"\n처음 10개 컬럼:")
    for i, col in enumerate(df_filtered.columns[:10]):
        print(f"  {i:2d}. {col}")
    
    print(f"\n처음 5행 데이터 (처음 6개 컬럼):")
    print(df_filtered.iloc[:5, :6])
    
    # openpyxl로 셀 병합 확인
    wb_filtered = load_workbook('output/anomaly_detection_filtered.xlsx')
    ws_filtered = wb_filtered[sheet_name]
    print(f"\n셀 병합 개수: {len(ws_filtered.merged_cells.ranges)}")
    print(f"처음 5개 병합 범위:")
    for i, mr in enumerate(list(ws_filtered.merged_cells.ranges)[:5]):
        print(f"  {i+1}. {mr}")
    
    # 차이점 분석
    print("\n" + "=" * 80)
    print("차이점 분석")
    print("=" * 80)
    
    # 1. 컬럼 수 차이
    print(f"\n1. 컬럼 수:")
    print(f"   원본: {len(df_original.columns)}")
    print(f"   필터링: {len(df_filtered.columns)}")
    print(f"   차이: {len(df_original.columns) - len(df_filtered.columns)} 컬럼 감소")
    
    # 2. 처음 3개 컬럼 비교
    print(f"\n2. 처음 3개 컬럼 비교:")
    print(f"   원본:    {df_original.columns[:3].tolist()}")
    print(f"   필터링:  {df_filtered.columns[:3].tolist()}")
    same = df_original.columns[:3].tolist() == df_filtered.columns[:3].tolist()
    print(f"   일치: {'✓' if same else '✗'}")
    
    # 3. PID 컬럼과 Unnamed 쌍 확인
    print(f"\n3. PID-Unnamed 쌍 패턴:")
    print(f"   원본 PID 컬럼:")
    orig_pids = [col for col in df_original.columns if col.startswith('PID_')]
    for i, pid in enumerate(orig_pids[:5]):
        idx = df_original.columns.tolist().index(pid)
        next_col = df_original.columns[idx+1] if idx+1 < len(df_original.columns) else "N/A"
        print(f"     {pid:40s} -> {next_col}")
    
    print(f"\n   필터링 PID 컬럼:")
    filt_pids = [col for col in df_filtered.columns if col.startswith('PID_')]
    for i, pid in enumerate(filt_pids[:5]):
        idx = df_filtered.columns.tolist().index(pid)
        next_col = df_filtered.columns[idx+1] if idx+1 < len(df_filtered.columns) else "N/A"
        print(f"     {pid:40s} -> {next_col}")
    
    # 4. 데이터 일치 확인 (처음 3행, 처음 3개 컬럼)
    print(f"\n4. 데이터 일치 확인 (처음 3행, 처음 3개 컬럼):")
    data_same = df_original.iloc[:3, :3].equals(df_filtered.iloc[:3, :3])
    print(f"   일치: {'✓' if data_same else '✗'}")
    if not data_same:
        print("\n   원본:")
        print(df_original.iloc[:3, :3])
        print("\n   필터링:")
        print(df_filtered.iloc[:3, :3])
    
    # 5. 셀 병합 패턴 확인
    print(f"\n5. 셀 병합:")
    print(f"   원본: {len(ws_original.merged_cells.ranges)} 범위")
    print(f"   필터링: {len(ws_filtered.merged_cells.ranges)} 범위")

if __name__ == '__main__':
    precise_comparison()
