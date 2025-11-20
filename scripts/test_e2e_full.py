#!/usr/bin/env python3
"""
전체 시스템 E2E 테스트

PCAP → 53-PID 파싱 → Excel 출력 전체 파이프라인 검증
"""

import sys
import os
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

from src.packet_source import PcapSource
from src.parser import EnhancedRTPSParser
from src.sink import DataFrameSink
from src.processor import TimeWindowProcessor
from src.transformer import ParticipantGrouper, PivotTableBuilder
from src.excel_writer import ExcelWriter
from openpyxl import load_workbook


def test_pcap_file(pcap_file: str, output_dir: str = 'output'):
    """
    단일 PCAP 파일 전체 파이프라인 테스트
    """
    print("\n" + "=" * 80)
    print(f"테스트: {Path(pcap_file).name}")
    print("=" * 80)
    
    if not Path(pcap_file).exists():
        print(f"❌ 파일 없음: {pcap_file}")
        return False
    
    try:
        # 1. 파싱
        print("[1/5] PCAP 파싱...")
        source = PcapSource(pcap_file)
        parser = EnhancedRTPSParser()
        sink = DataFrameSink()
        processor = TimeWindowProcessor(window_seconds=1.0)
        
        processor.process_stream(source, parser, sink)
        df = sink.get_result()
        
        print(f"  ✓ {len(df)}개 submessage")
        
        # PID 필드 확인
        if 'pids' in df.columns:
            all_pid_keys = set()
            for pids in df['pids'].dropna():
                if isinstance(pids, dict):
                    all_pid_keys.update(pids.keys())
            print(f"  ✓ {len(all_pid_keys)}개 고유 PID 필드")
        else:
            print("  ⚠️  pids 컬럼 없음")
            return False
        
        # 2. 그룹화
        print("[2/5] 참여자 그룹화...")
        submessages = df.to_dict('records')
        grouper = ParticipantGrouper()
        grouped = grouper.group_by_participant(submessages)
        
        print(f"  ✓ {len(grouped)}개 참여자")
        
        if len(grouped) == 0:
            print("  ⚠️  참여자 없음")
            return False
        
        # 3. Pivot 생성
        print("[3/5] Pivot 테이블 생성...")
        builder = PivotTableBuilder(window_size=1.0)
        
        participant_dataframes = {}
        for idx, (participant_id, messages) in enumerate(grouped.items(), start=1):
            hostId, appId, instanceId = participant_id
            guid_str = f"{hostId:08x}.{appId:08x}.{instanceId:08x}"
            
            pivot_df = builder.build(messages, participant_id=participant_id)
            
            # 시트명 생성 (Excel 31자 제한 고려)
            if len(f"Node_{idx}_{guid_str}") <= 31:
                node_name = f"Node_{idx}_{guid_str}"
            else:
                node_name = f"Node_{idx}_{hostId:08x}..{instanceId:08x}"
            
            participant_dataframes[node_name] = pivot_df
        
        # Overview 생성
        import pandas as pd
        overview_data = []
        for idx, (participant_id, messages) in enumerate(grouped.items(), start=1):
            hostId, appId, instanceId = participant_id
            guid_str = f"{hostId:08x}.{appId:08x}.{instanceId:08x}"
            
            if len(f"Node_{idx}_{guid_str}") <= 31:
                sheet_name = f"Node_{idx}_{guid_str}"
            else:
                sheet_name = f"Node_{idx}_{hostId:08x}..{instanceId:08x}"
            
            overview_data.append({
                'Node': f"Node_{idx}",
                'Sheet Name': sheet_name,
                'Participant ID': guid_str,
                'Submessages': len(messages),
                'Rows': len(participant_dataframes[sheet_name])
            })
        
        overview_df = pd.DataFrame(overview_data)
        
        print(f"  ✓ {len(participant_dataframes)}개 노드 시트")
        print(f"  ✓ Overview: {overview_df.shape}")
        
        # 4. Excel 생성
        print("[4/5] Excel 파일 생성...")
        Path(output_dir).mkdir(exist_ok=True)
        
        pcap_name = Path(pcap_file).stem
        output_file = Path(output_dir) / f"{pcap_name}_analysis.xlsx"
        
        writer = ExcelWriter(output_path=output_file)
        writer.write_overview(overview_df)
        writer.write_participant_sheets(participant_dataframes)
        saved_path = writer.save()
        writer.close()
        
        file_size = os.path.getsize(saved_path)
        print(f"  ✓ {saved_path}")
        print(f"  ✓ {file_size:,} bytes")
        
        # 5. Excel 검증
        print("[5/5] Excel 검증...")
        wb = load_workbook(str(saved_path))
        
        try:
            print(f"  ✓ {len(wb.sheetnames)}개 시트")
            
            # 첫 번째 노드 시트 검증
            OVERVIEW_SHEET_INDEX = 0
            FIRST_NODE_SHEET_INDEX = 1
            
            if len(wb.sheetnames) > FIRST_NODE_SHEET_INDEX:
                node_sheet = wb[wb.sheetnames[FIRST_NODE_SHEET_INDEX]]
                print(f"  ✓ 노드 시트: {node_sheet.max_row}행 × {node_sheet.max_column}열")
                
                # 병합된 셀 확인
                merged = len(list(node_sheet.merged_cells.ranges))
                print(f"  ✓ {merged}개 병합된 셀 (헤더)")
            
            # 데이터 무결성 검증: 원본 submessage 수 == Excel 총 행 수
            total_excel_rows = sum(
                wb[sheet].max_row - 2  # 헤더 2행 제외
                for sheet in wb.sheetnames[FIRST_NODE_SHEET_INDEX:]  # Overview 제외
            )
            
            if len(df) == total_excel_rows:
                print(f"  ✓ 데이터 무결성: {len(df)} submessages = {total_excel_rows} Excel rows")
            else:
                print(f"  ⚠️  데이터 불일치: {len(df)} submessages != {total_excel_rows} Excel rows")
        
        finally:
            wb.close()
        
        print("✅ 성공!")
        return True
        
    except Exception as e:
        print(f"❌ 오류: {e}")
        import traceback
        traceback.print_exc()
        return False


def main():
    print("=" * 80)
    print("전체 시스템 E2E 테스트")
    print("53-PID 시스템 통합 검증")
    print("=" * 80)
    
    # 테스트할 PCAP 파일들
    pcap_files = [
        'data/participated_57_first_bye.pcapng',
    ]
    
    results = {}
    for pcap_file in pcap_files:
        result = test_pcap_file(pcap_file)
        results[pcap_file] = result
    
    # 최종 결과
    print("\n" + "=" * 80)
    print("최종 결과")
    print("=" * 80)
    
    total = len(results)
    passed = sum(1 for r in results.values() if r)
    
    for pcap, result in results.items():
        status = "✅ PASS" if result else "❌ FAIL"
        print(f"{status} - {Path(pcap).name}")
    
    print("=" * 80)
    print(f"총 {total}개 테스트 중 {passed}개 통과")
    
    if passed == total:
        print("🎉 모든 테스트 통과!")
        print("=" * 80)
        return 0
    else:
        print(f"⚠️  {total - passed}개 테스트 실패")
        print("=" * 80)
        return 1


if __name__ == '__main__':
    sys.exit(main())
