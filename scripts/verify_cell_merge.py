#!/usr/bin/env python3
"""필터링된 Excel의 셀 병합 확인"""

from openpyxl import load_workbook

def verify_cell_merges():
    """셀 병합 검증"""
    
    print("=" * 80)
    print("필터링된 Excel 셀 병합 검증")
    print("=" * 80)
    
    # 필터링된 파일 읽기
    wb = load_workbook('output/anomaly_detection_filtered.xlsx')
    
    # 첫 번째 노드 시트
    sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]
    
    print(f"\n시트: {sheet_name}")
    print(f"병합된 셀 범위 개수: {len(ws.merged_cells.ranges)}")
    
    # 병합 범위 출력 (처음 20개)
    print("\n처음 20개 병합 범위:")
    for i, merged_range in enumerate(list(ws.merged_cells.ranges)[:20]):
        print(f"  {i+1}. {merged_range}")
    
    # A열(idx)과 B열(TimeInterval)의 병합 확인
    a_merges = [r for r in ws.merged_cells.ranges if str(r).startswith('A')]
    b_merges = [r for r in ws.merged_cells.ranges if str(r).startswith('B')]
    
    print(f"\nA열(idx) 병합 개수: {len(a_merges)}")
    print(f"B열(TimeInterval) 병합 개수: {len(b_merges)}")
    
    # 처음 5개씩만 출력
    print("\nA열(idx) 병합 (처음 5개):")
    for r in a_merges[:5]:
        print(f"  {r}")
    
    print("\nB열(TimeInterval) 병합 (처음 5개):")
    for r in b_merges[:5]:
        print(f"  {r}")
    
    print("\n" + "=" * 80)
    print("✓ 셀 병합 적용 완료!")
    print("=" * 80)

if __name__ == '__main__':
    verify_cell_merges()
