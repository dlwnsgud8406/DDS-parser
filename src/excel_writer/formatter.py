#!/usr/bin/env python3
"""
Excel 컬럼 포맷팅

컬럼 너비 자동 조정, 값 포맷팅 등
"""

from openpyxl.utils import get_column_letter


class ColumnFormatter:
    """Excel 컬럼 포맷팅 관리"""
    
    @staticmethod
    def auto_adjust_column_width(worksheet, min_width=10, max_width=50):
        """
        컬럼 너비 자동 조정
        
        Args:
            worksheet: openpyxl Worksheet 객체
            min_width: 최소 너비
            max_width: 최대 너비
        """
        # 컬럼별로 처리 (병합된 셀 무시)
        for col_idx in range(1, worksheet.max_column + 1):
            length = 0
            column_letter = get_column_letter(col_idx)
            
            for row_idx in range(1, worksheet.max_row + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                
                # 병합된 셀은 스킵
                if isinstance(cell, type(cell)) and hasattr(cell, 'value'):
                    try:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > length:
                                length = cell_length
                    except:
                        pass
            
            # 너비 계산 (1.2배 여유 공간)
            adjusted_width = min(max(length * 1.2, min_width), max_width)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    @staticmethod
    def format_timestamp(value):
        """
        Timestamp 포맷팅 (소수점 6자리)
        
        Args:
            value: Timestamp 값 (float 또는 str)
        
        Returns:
            str: 포맷팅된 문자열 (예: "0.123456")
        """
        if value is None or value == '':
            return ''
        
        try:
            # float로 변환
            if isinstance(value, str):
                # "0.123456789" → 0.123456789
                value = float(value)
            
            # 소수점 6자리로 포맷
            return f"{value:.6f}"
        except (ValueError, TypeError):
            return str(value)
    
    @staticmethod
    def format_hex_value(value):
        """
        16진수 값 포맷팅 (0x 접두사 유지)
        
        Args:
            value: 16진수 값 (int, str, 또는 0x로 시작하는 문자열)
        
        Returns:
            str: 포맷팅된 문자열 (예: "0x0000001f")
        """
        if value is None or value == '':
            return ''
        
        try:
            # 이미 0x로 시작하는 문자열이면 그대로 반환
            if isinstance(value, str) and value.startswith('0x'):
                return value
            
            # 문자열인 경우 int로 변환
            if isinstance(value, str):
                # "31" 또는 "0x1f" → int
                value = int(value, 0)
            
            # int인 경우 0x 접두사 포맷
            if isinstance(value, int):
                return f"0x{value:08x}"
            
            return str(value)
        except (ValueError, TypeError):
            return str(value)
    
    @staticmethod
    def format_time_interval(value):
        """
        TimeInterval 포맷팅
        
        Args:
            value: TimeInterval 값 (예: "0.0-1.0s")
        
        Returns:
            str: 그대로 반환 (이미 포맷팅됨)
        """
        return str(value) if value else ''
    
    @staticmethod
    def apply_number_format(cell, column_name):
        """
        셀에 숫자 포맷 적용
        
        Args:
            cell: openpyxl Cell 객체
            column_name: 컬럼 이름
        """
        if 'Timestamp' in column_name:
            # Timestamp: 소수점 6자리
            cell.number_format = '0.000000'
        elif 'Value' in column_name:
            # Hex Value: 텍스트 형식 (0x 접두사 유지)
            cell.number_format = '@'
        elif column_name == 'idx':
            # idx: 정수
            cell.number_format = '0'
