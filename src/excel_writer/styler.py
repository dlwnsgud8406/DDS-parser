#!/usr/bin/env python3
"""
Excel 셀 스타일링

셀 스타일, 폰트, 배경색, 정렬 등을 관리
"""

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ExcelStyler:
    """Excel 셀 스타일링 관리"""
    
    # 색상 정의
    HEADER_BG_COLOR = "4472C4"  # 파란색
    EVEN_ROW_BG_COLOR = "F2F2F2"  # 연한 회색
    ODD_ROW_BG_COLOR = "FFFFFF"  # 흰색
    
    # 폰트 정의
    HEADER_FONT = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
    DATA_FONT = Font(name='Calibri', size=10)
    
    # 정렬 정의
    CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
    LEFT_ALIGN = Alignment(horizontal='left', vertical='center')
    RIGHT_ALIGN = Alignment(horizontal='right', vertical='center')
    
    # 테두리 정의
    THIN_BORDER = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    @classmethod
    def apply_header_style(cls, cell, merge_cell=False):
        """
        헤더 스타일 적용
        
        Args:
            cell: openpyxl Cell 객체
            merge_cell: 병합 셀 여부 (병합 셀은 테두리만 적용)
        """
        cell.font = cls.HEADER_FONT
        cell.fill = PatternFill(start_color=cls.HEADER_BG_COLOR, 
                                end_color=cls.HEADER_BG_COLOR, 
                                fill_type='solid')
        cell.alignment = cls.CENTER_ALIGN
        cell.border = cls.THIN_BORDER
    
    @classmethod
    def apply_data_style(cls, cell, is_even_row=False, align='left'):
        """
        데이터 셀 스타일 적용
        
        Args:
            cell: openpyxl Cell 객체
            is_even_row: 짝수 행 여부 (배경색 번갈아가며)
            align: 정렬 방식 ('left', 'center', 'right')
        """
        cell.font = cls.DATA_FONT
        
        # 배경색 (번갈아가며)
        bg_color = cls.EVEN_ROW_BG_COLOR if is_even_row else cls.ODD_ROW_BG_COLOR
        cell.fill = PatternFill(start_color=bg_color, 
                                end_color=bg_color, 
                                fill_type='solid')
        
        # 정렬
        if align == 'center':
            cell.alignment = cls.CENTER_ALIGN
        elif align == 'right':
            cell.alignment = cls.RIGHT_ALIGN
        else:
            cell.alignment = cls.LEFT_ALIGN
        
        cell.border = cls.THIN_BORDER
    
    @classmethod
    def merge_and_style_header(cls, worksheet, start_col, end_col, row, text):
        """
        헤더 셀 수평 병합 및 스타일 적용
        
        Args:
            worksheet: openpyxl Worksheet 객체
            start_col: 시작 컬럼 인덱스 (1-based)
            end_col: 끝 컬럼 인덱스 (1-based)
            row: 행 번호 (1-based)
            text: 셀 텍스트
        """
        # 셀 병합
        start_col_letter = get_column_letter(start_col)
        end_col_letter = get_column_letter(end_col)
        
        if start_col == end_col:
            # 병합 없음
            cell = worksheet.cell(row=row, column=start_col)
            cell.value = text
            cls.apply_header_style(cell)
        else:
            # 병합
            merge_range = f"{start_col_letter}{row}:{end_col_letter}{row}"
            worksheet.merge_cells(merge_range)
            
            # 병합된 첫 번째 셀에 값 설정
            cell = worksheet.cell(row=row, column=start_col)
            cell.value = text
            cls.apply_header_style(cell)
            
            # 병합된 나머지 셀에도 스타일 적용 (값은 없음)
            for col in range(start_col + 1, end_col + 1):
                cell = worksheet.cell(row=row, column=col)
                cls.apply_header_style(cell, merge_cell=True)
    
    @classmethod
    def merge_and_style_header_vertical(cls, worksheet, col, start_row, end_row, text):
        """
        헤더 셀 수직 병합 및 스타일 적용
        
        Args:
            worksheet: openpyxl Worksheet 객체
            col: 컬럼 인덱스 (1-based)
            start_row: 시작 행 번호 (1-based)
            end_row: 끝 행 번호 (1-based)
            text: 셀 텍스트
        """
        col_letter = get_column_letter(col)
        
        if start_row == end_row:
            # 병합 없음
            cell = worksheet.cell(row=start_row, column=col)
            cell.value = text
            cls.apply_header_style(cell)
        else:
            # 수직 병합
            merge_range = f"{col_letter}{start_row}:{col_letter}{end_row}"
            worksheet.merge_cells(merge_range)
            
            # 병합된 첫 번째 셀에 값 설정
            cell = worksheet.cell(row=start_row, column=col)
            cell.value = text
            cls.apply_header_style(cell)
            
            # 병합된 나머지 셀에도 스타일 적용 (값은 없음)
            for row in range(start_row + 1, end_row + 1):
                cell = worksheet.cell(row=row, column=col)
                cls.apply_header_style(cell, merge_cell=True)
    
    @classmethod
    def apply_column_styles(cls, worksheet, start_row, data_rows):
        """
        전체 데이터 영역에 스타일 적용
        
        Args:
            worksheet: openpyxl Worksheet 객체
            start_row: 데이터 시작 행 (헤더 다음)
            data_rows: 데이터 행 수
        """
        max_col = worksheet.max_column
        
        for row_idx in range(start_row, start_row + data_rows):
            is_even = (row_idx - start_row) % 2 == 0
            
            for col_idx in range(1, max_col + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                
                # 컬럼별 정렬 결정
                col_letter = get_column_letter(col_idx)
                col_name = worksheet.cell(row=start_row - 1, column=col_idx).value
                
                if col_name:
                    if 'TimeInterval' in str(col_name):
                        align = 'center'
                    elif 'Timestamp' in str(col_name) or 'Value' in str(col_name):
                        align = 'right'
                    else:
                        align = 'center' if col_idx == 1 else 'left'
                else:
                    align = 'left'
                
                cls.apply_data_style(cell, is_even_row=is_even, align=align)
