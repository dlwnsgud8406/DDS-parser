#!/usr/bin/env python3
"""
Excel 파일 생성

DataFrame을 Excel 파일로 변환
"""

from datetime import datetime
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from .styler import ExcelStyler
from .formatter import ColumnFormatter


class ExcelWriter:
    """Excel 파일 생성 및 관리"""
    
    def __init__(self, output_path=None):
        """
        Args:
            output_path: 출력 파일 경로 (None이면 자동 생성)
        """
        if output_path is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = f"rtps_analysis_{timestamp}.xlsx"
        
        self.output_path = Path(output_path)
        self.workbook = Workbook()
        
        # 기본 시트 삭제
        if 'Sheet' in self.workbook.sheetnames:
            del self.workbook['Sheet']
    
    def write_overview(self, df_overview):
        """
        Overview 시트 생성 (Sheet 1)
        
        Args:
            df_overview: pandas.DataFrame
                Columns: Node | HostId | AppId | MessageCount | EntityCount | TimeSpan | SubmsgTypes
        """
        # 시트 생성
        ws = self.workbook.create_sheet("Overview", 0)
        
        # DataFrame을 시트에 쓰기
        self._write_dataframe_to_sheet(ws, df_overview, sheet_type='overview')
        
        print(f"  ✓ Overview 시트 생성: {len(df_overview)} nodes")
    
    def write_qos_summary(self, df_qos):
        """
        QoS Summary 시트 생성
        
        Args:
            df_qos: pandas.DataFrame
                Columns: Topic | Reliability | Durability | Frequency | Messages
        """
        # 시트 생성 (Overview 다음)
        ws = self.workbook.create_sheet("QoS_Summary", 1)
        
        # DataFrame을 시트에 쓰기
        self._write_dataframe_to_sheet(ws, df_qos, sheet_type='overview')
        
        print(f"  ✓ QoS Summary 시트 생성: {len(df_qos)} topics")
    
    def write_sedp_sheet(self, df_sedp):
        """
        SEDP (Endpoint Discovery) 시트 생성
        
        Args:
            df_sedp: pandas.DataFrame
                Columns: endpoint_kind | hostId | appId | instanceId | entityId | 
                         topic | type | reliability | durability | ...
        """
        # 시트 생성 (QoS Summary 다음)
        ws = self.workbook.create_sheet("SEDP_Endpoints", 2)
        
        # DataFrame을 시트에 쓰기
        self._write_dataframe_to_sheet(ws, df_sedp, sheet_type='overview')
        
        print(f"  ✓ SEDP 시트 생성: {len(df_sedp)} endpoints")
    
    def write_topic_sheets(self, topic_dataframes, max_rows=None):
        """
        각 Topic별 시트 생성
        
        Args:
            topic_dataframes: dict
                {topic_sheet_name: DataFrame, ...}
                DataFrame columns: idx | TimeInterval | PID_XXX_Timestamp | PID_XXX_Value | ...
            max_rows: 최대 행 수 제한 (None이면 제한 없음)
        """
        for sheet_name, df in topic_dataframes.items():
            # 시트명이 31자를 초과하면 잘라냄 (Excel 제한)
            sheet_name = sheet_name[:31] if len(sheet_name) > 31 else sheet_name
            
            # 대용량 데이터 경고
            if max_rows and len(df) > max_rows:
                print(f"  ⚠️  {sheet_name}: {len(df):,}행 (상위 {max_rows:,}행만 출력)")
                df = df.head(max_rows)
            
            # 시트 생성
            ws = self.workbook.create_sheet(sheet_name)
            
            # DataFrame을 시트에 쓰기 (PID 헤더 병합 포함)
            self._write_dataframe_to_sheet(ws, df, sheet_type='participant')
            
            print(f"  ✓ {sheet_name}: {df.shape}")
    
    def write_participant_sheets(self, participant_dataframes, max_rows=None):
        """
        각 Participant별 시트 생성 (Sheet 2~N)
        
        Args:
            participant_dataframes: dict
                {node_name: DataFrame, ...}
                DataFrame columns: idx | TimeInterval | PID_XXX_Timestamp | PID_XXX_Value | ...
            max_rows: 최대 행 수 제한 (None이면 제한 없음)
        """
        for node_name, df in participant_dataframes.items():
            # 시트명이 31자를 초과하면 잘라냄 (Excel 제한)
            sheet_name = node_name[:31] if len(node_name) > 31 else node_name
            
            # 대용량 데이터 경고
            if max_rows and len(df) > max_rows:
                print(f"  ⚠️  {sheet_name}: {len(df):,}행 (상위 {max_rows:,}행만 출력)")
                df = df.head(max_rows)
            
            # 시트 생성
            ws = self.workbook.create_sheet(sheet_name)
            
            # DataFrame을 시트에 쓰기 (PID 헤더 병합 포함)
            self._write_dataframe_to_sheet(ws, df, sheet_type='participant')
            
            print(f"  ✓ {sheet_name}: {df.shape}")
    
    def write_node_sheets(self, node_dataframes, max_rows=None):
        """
        각 ROS2 노드별 시트 생성
        
        Args:
            node_dataframes: dict
                {node_name: DataFrame, ...}
                DataFrame columns: idx | TimeInterval | topic | submsg_name | PID_XXX | ...
            max_rows: 최대 행 수 제한 (None이면 제한 없음)
        """
        for node_name, df in node_dataframes.items():
            # 시트명이 31자를 초과하면 잘라냄 (Excel 제한)
            sheet_name = node_name[:31] if len(node_name) > 31 else node_name
            
            # 대용량 데이터 경고
            if max_rows and len(df) > max_rows:
                print(f"  ⚠️  {sheet_name}: {len(df):,}행 (상위 {max_rows:,}행만 출력)")
                df = df.head(max_rows)
            
            # 시트 생성
            ws = self.workbook.create_sheet(sheet_name)
            
            # DataFrame을 시트에 쓰기 (PID 헤더 병합 포함)
            self._write_dataframe_to_sheet(ws, df, sheet_type='participant')
            
            print(f"  ✓ {sheet_name}: {df.shape}")
    
    def _write_dataframe_to_sheet(self, worksheet, df, sheet_type='participant'):
        """
        DataFrame을 시트에 쓰기 (헤더 포함)
        
        Args:
            worksheet: openpyxl Worksheet 객체
            df: pandas.DataFrame
            sheet_type: 'overview' 또는 'participant'
        """
        # DataFrame을 행으로 변환
        rows = list(dataframe_to_rows(df, index=False, header=True))
        
        if sheet_type == 'participant':
            # Participant 시트: PID 헤더 2행 구조
            self._write_participant_sheet(worksheet, df, rows)
        else:
            # Overview 시트: 단순 테이블
            self._write_simple_sheet(worksheet, rows)
    
    def _write_simple_sheet(self, worksheet, rows):
        """
        단순 테이블 시트 작성 (Overview)
        
        Args:
            worksheet: openpyxl Worksheet 객체
            rows: DataFrame rows (헤더 포함)
        """
        # 데이터 쓰기
        for row_idx, row_data in enumerate(rows, start=1):
            for col_idx, value in enumerate(row_data, start=1):
                cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
                
                if row_idx == 1:
                    # 헤더
                    ExcelStyler.apply_header_style(cell)
                else:
                    # 데이터
                    is_even = (row_idx - 2) % 2 == 0
                    ExcelStyler.apply_data_style(cell, is_even_row=is_even, align='left')
        
        # 컬럼 너비 자동 조정
        ColumnFormatter.auto_adjust_column_width(worksheet)
    
    def _write_participant_sheet(self, worksheet, df, rows):
        """
        Participant 시트 작성 (PID 헤더 2행 구조 + TimeInterval 셀 병합)
        
        Args:
            worksheet: openpyxl Worksheet 객체
            df: pandas.DataFrame
            rows: DataFrame rows (헤더 포함)
        """
        # 1행: PID 헤더 (병합)
        # 2행: Timestamp/Value 서브헤더
        
        # 컬럼 분석
        columns = df.columns.tolist()
        header_groups = self._analyze_pid_columns(columns)
        
        # 1행: PID 헤더 작성 (병합)
        current_col = 1
        for group in header_groups:
            if group['type'] == 'basic':
                # 기본 컬럼 (idx, TimeInterval) - 수직 병합 (1행~2행)
                worksheet.cell(row=1, column=current_col, value=group['name'])
                ExcelStyler.merge_and_style_header_vertical(
                    worksheet, current_col, 1, 2, group['name']
                )
                current_col += 1
            else:
                # PID 컬럼 (Timestamp/Value 쌍) - 수평 병합 (1행)
                ExcelStyler.merge_and_style_header(
                    worksheet, current_col, current_col + 1, 1, group['pid_name']
                )
                
                # 2행: Timestamp, Value
                ExcelStyler.merge_and_style_header(
                    worksheet, current_col, current_col, 2, 'Timestamp'
                )
                ExcelStyler.merge_and_style_header(
                    worksheet, current_col + 1, current_col + 1, 2, 'Value'
                )
                
                current_col += 2
        
        # 3행부터: 데이터 쓰기 (스타일 제거로 속도 향상)
        total_rows = len(rows) - 1  # 헤더 제외
        for data_idx, row_data in enumerate(rows[1:], start=3):  # rows[1:]은 헤더 제외
            # 진행 상황 표시 (1000행마다)
            if (data_idx - 3) % 1000 == 0:
                progress = ((data_idx - 3) / total_rows) * 100
                print(f"    진행: {data_idx - 3:,}/{total_rows:,} 행 ({progress:.1f}%)", end='\r')
            
            for col_idx, value in enumerate(row_data, start=1):
                # 데이터만 쓰기 (스타일 없음 - 속도 향상)
                worksheet.cell(row=data_idx, column=col_idx, value=value)
        
        print(f"\n    ✓ {len(rows)-1:,}행 작성 완료")
        
        # TimeInterval 셀 병합 처리
        if hasattr(df, 'attrs') and 'merge_info' in df.attrs:
            print(f"    셀 병합 처리 중...")
            merge_info = df.attrs['merge_info']
            self._merge_interval_cells(worksheet, merge_info, columns)
            print(f"    ✓ {len(merge_info):,}개 구간 병합 완료")
        
        # 컬럼 너비 자동 조정
        ColumnFormatter.auto_adjust_column_width(worksheet)
    
    def _merge_interval_cells(self, worksheet, merge_info, columns):
        """
        TimeInterval과 idx 컬럼의 셀 병합 처리
        
        Args:
            worksheet: openpyxl Worksheet 객체
            merge_info: [(start_row, end_row, window_key), ...]
            columns: 컬럼 리스트
        """
        from openpyxl.styles import Alignment
        
        # idx와 TimeInterval의 컬럼 인덱스 찾기
        try:
            idx_col = columns.index('idx') + 1
            interval_col = columns.index('TimeInterval') + 1
        except ValueError:
            return  # 컬럼이 없으면 스킵
        
        total_merges = len(merge_info)
        for merge_idx, (start_row, end_row, window_key) in enumerate(merge_info):
            # 진행 상황 표시 (1000개마다)
            if merge_idx % 1000 == 0:
                progress = (merge_idx / total_merges) * 100
                print(f"      병합: {merge_idx:,}/{total_merges:,} ({progress:.1f}%)", end='\r')
            
            # DataFrame의 행 인덱스를 Excel 행 번호로 변환 (헤더 2행 + 데이터 시작 행 3)
            excel_start_row = start_row + 3
            excel_end_row = end_row + 3
            
            # idx 컬럼 병합
            if excel_start_row < excel_end_row:
                worksheet.merge_cells(
                    start_row=excel_start_row,
                    start_column=idx_col,
                    end_row=excel_end_row,
                    end_column=idx_col
                )
                
                # 병합된 셀 스타일 (중앙 정렬)
                merged_cell = worksheet.cell(row=excel_start_row, column=idx_col)
                merged_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # TimeInterval 컬럼 병합
            if excel_start_row < excel_end_row:
                worksheet.merge_cells(
                    start_row=excel_start_row,
                    start_column=interval_col,
                    end_row=excel_end_row,
                    end_column=interval_col
                )
                
                # 병합된 셀 스타일 (중앙 정렬)
                merged_cell = worksheet.cell(row=excel_start_row, column=interval_col)
                merged_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        print()  # 진행률 표시 후 줄바꿈
    
    def _analyze_pid_columns(self, columns):
        """
        컬럼 분석하여 PID 그룹화
        
        Args:
            columns: 컬럼 리스트
        
        Returns:
            list: [
                {'type': 'basic', 'name': 'idx'},
                {'type': 'basic', 'name': 'TimeInterval'},
                {'type': 'pid', 'pid_name': 'PID_0x0006', 'columns': ['PID_0x0006_Timestamp', 'PID_0x0006_Value']},
                ...
            ]
        """
        groups = []
        i = 0
        
        while i < len(columns):
            col_name = columns[i]
            
            if col_name in ['idx', 'TimeInterval']:
                # 기본 컬럼
                groups.append({'type': 'basic', 'name': col_name})
                i += 1
            elif '_Timestamp' in col_name:
                # PID 컬럼 시작 (Timestamp/Value 쌍)
                pid_name = col_name.replace('_Timestamp', '')
                
                # Value 컬럼도 있는지 확인
                if i + 1 < len(columns) and columns[i + 1] == f"{pid_name}_Value":
                    groups.append({
                        'type': 'pid',
                        'pid_name': pid_name,
                        'columns': [col_name, columns[i + 1]]
                    })
                    i += 2
                else:
                    # Value 컬럼이 없으면 단독 처리
                    groups.append({'type': 'basic', 'name': col_name})
                    i += 1
            else:
                # 기타 컬럼
                groups.append({'type': 'basic', 'name': col_name})
                i += 1
        
        return groups
    
    def save(self):
        """
        Excel 파일 저장
        
        Returns:
            Path: 저장된 파일 경로
        """
        self.workbook.save(self.output_path)
        return self.output_path
    
    def close(self):
        """Workbook 닫기"""
        self.workbook.close()
