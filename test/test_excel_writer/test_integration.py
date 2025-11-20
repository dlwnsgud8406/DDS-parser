#!/usr/bin/env python3
"""
Excel Writer 통합 테스트
전체 파이프라인: PCAP → Parse → Transform → Excel
"""

import pytest
import sys
from pathlib import Path
import os

# 프로젝트 루트 추가
sys.path.insert(0, str(Path(__file__).parent.parent.parent))

from src.parser import EnhancedRTPSParser
from src.packet_source import PcapSource
from src.transformer import ParticipantGrouper, PivotTableBuilder
from src.excel_writer import ExcelWriter


class TestExcelWriterIntegration:
    """Excel Writer 통합 테스트"""
    
    @pytest.fixture
    def sample_pcap(self):
        """샘플 PCAP 경로"""
        pcap_path = Path(__file__).parent.parent.parent / "data" / "participated_57_first_bye.pcapng"
        if not pcap_path.exists():
            pytest.skip(f"PCAP 파일 없음: {pcap_path}")
        return str(pcap_path)
    
    @pytest.fixture
    def output_dir(self):
        """출력 디렉토리"""
        output_dir = Path(__file__).parent.parent.parent / "output"
        output_dir.mkdir(exist_ok=True)
        return output_dir
    
    @pytest.fixture
    def parsed_and_transformed_data(self, sample_pcap):
        """파싱 및 변환된 데이터"""
        print(f"\n{'='*70}")
        print("전체 파이프라인 테스트: PCAP → Parse → Transform → Excel")
        print(f"{'='*70}\n")
        
        # 1. Parse
        print("1. 파싱 중...")
        parser = EnhancedRTPSParser()
        source = PcapSource(sample_pcap)
        packets = list(source)
        submessages = parser.parse_batch(packets)
        print(f"   ✓ {len(packets)} packets → {len(submessages)} submessages")
        
        # 2. Transform
        print("\n2. 데이터 변환 중...")
        grouper = ParticipantGrouper()
        grouped = grouper.group_by_participant(submessages)
        print(f"   ✓ {len(grouped)} participants")
        
        builder = PivotTableBuilder(window_size=1.0)
        
        participant_dataframes = {}
        for idx, (participant_id, messages) in enumerate(grouped.items(), start=1):
            df = builder.build(messages, participant_id=participant_id)
            node_name = grouper.format_participant_name(participant_id, index=idx)
            participant_dataframes[node_name] = df
            print(f"   ✓ {node_name}: {df.shape}")
        
        overview_df = builder.build_summary(grouped)
        print(f"   ✓ Overview: {overview_df.shape}")
        
        return {
            'participant_dfs': participant_dataframes,
            'overview_df': overview_df
        }
    
    def test_excel_writer_basic(self, parsed_and_transformed_data, output_dir):
        """Excel Writer 기본 동작 테스트"""
        data = parsed_and_transformed_data
        
        print(f"\n3. Excel 파일 생성 중...")
        
        # Excel Writer 생성
        output_path = output_dir / "test_rtps_analysis.xlsx"
        writer = ExcelWriter(output_path=str(output_path))
        
        # Overview 시트 작성
        writer.write_overview(data['overview_df'])
        
        # Participant 시트 작성
        writer.write_participant_sheets(data['participant_dfs'])
        
        # 저장
        saved_path = writer.save()
        writer.close()
        
        print(f"   ✓ Excel 파일 저장: {saved_path}")
        
        # 검증
        assert saved_path.exists(), f"파일이 생성되지 않음: {saved_path}"
        assert saved_path.stat().st_size > 0, "파일 크기가 0"
        
        print(f"\n{'='*70}")
        print(f"✅ Excel 파일 생성 성공!")
        print(f"   파일 위치: {saved_path}")
        print(f"   파일 크기: {saved_path.stat().st_size / 1024:.2f} KB")
        print(f"{'='*70}")
    
    def test_excel_structure(self, parsed_and_transformed_data, output_dir):
        """Excel 파일 구조 검증"""
        from openpyxl import load_workbook
        
        data = parsed_and_transformed_data
        
        # Excel 파일 생성
        output_path = output_dir / "test_rtps_structure.xlsx"
        writer = ExcelWriter(output_path=str(output_path))
        writer.write_overview(data['overview_df'])
        writer.write_participant_sheets(data['participant_dfs'])
        saved_path = writer.save()
        writer.close()
        
        # 파일 로드
        wb = load_workbook(saved_path)
        
        print(f"\n시트 구조 검증:")
        print(f"  - 총 시트 수: {len(wb.sheetnames)}")
        
        # Overview 시트 확인
        assert 'Overview' in wb.sheetnames, "Overview 시트 없음"
        overview_ws = wb['Overview']
        print(f"  - Overview: {overview_ws.max_row} rows, {overview_ws.max_column} columns")
        
        # Participant 시트 확인
        participant_count = len(data['participant_dfs'])
        assert len(wb.sheetnames) == participant_count + 1, \
            f"시트 개수 불일치: expected {participant_count + 1}, got {len(wb.sheetnames)}"
        
        for sheet_name in wb.sheetnames[1:]:  # Overview 제외
            ws = wb[sheet_name]
            print(f"  - {sheet_name}: {ws.max_row} rows, {ws.max_column} columns")
            
            # 첫 번째 셀 확인 (헤더)
            assert ws.cell(1, 1).value is not None, f"{sheet_name}: 헤더가 비어있음"
        
        wb.close()
        
        print(f"\n✅ Excel 구조 검증 완료")
    
    def test_full_pipeline_with_output(self, sample_pcap, output_dir):
        """전체 파이프라인 테스트 (출력 포함)"""
        print(f"\n{'='*70}")
        print("전체 파이프라인 E2E 테스트")
        print(f"{'='*70}\n")
        
        # 1. Parse
        print("1. 파싱 중...")
        parser = EnhancedRTPSParser()
        source = PcapSource(sample_pcap)
        packets = list(source)
        submessages = parser.parse_batch(packets)
        print(f"   ✓ {len(packets)} packets → {len(submessages)} submessages")
        
        # 2. Group by Participant
        print("\n2. Participant별 그룹화 중...")
        grouper = ParticipantGrouper()
        grouped = grouper.group_by_participant(submessages)
        print(f"   ✓ {len(grouped)} participants")
        
        # 3. Transform to DataFrames
        print("\n3. DataFrame 변환 중...")
        builder = PivotTableBuilder(window_size=1.0)
        
        participant_dataframes = {}
        for idx, (participant_id, messages) in enumerate(grouped.items(), start=1):
            df = builder.build(messages, participant_id=participant_id)
            node_name = grouper.format_participant_name(participant_id, index=idx)
            participant_dataframes[node_name] = df
            print(f"   ✓ {node_name}: {df.shape}")
        
        overview_df = builder.build_summary(grouped)
        print(f"   ✓ Overview: {overview_df.shape}")
        
        # 4. Write to Excel
        print("\n4. Excel 파일 생성 중...")
        output_path = output_dir / "rtps_analysis_full_pipeline.xlsx"
        
        writer = ExcelWriter(output_path=str(output_path))
        writer.write_overview(overview_df)
        writer.write_participant_sheets(participant_dataframes)
        saved_path = writer.save()
        writer.close()
        
        print(f"   ✓ 저장 완료: {saved_path}")
        
        # 검증
        assert saved_path.exists()
        assert saved_path.stat().st_size > 0
        
        print(f"\n{'='*70}")
        print(f"✅ 전체 파이프라인 성공!")
        print(f"   📊 Excel 파일: {saved_path}")
        print(f"   📏 파일 크기: {saved_path.stat().st_size / 1024:.2f} KB")
        print(f"   📄 시트 개수: {len(participant_dataframes) + 1}")
        print(f"   🗂️  Participants: {len(participant_dataframes)}")
        print(f"{'='*70}")


# pytest 실행용
if __name__ == "__main__":
    pytest.main([__file__, "-v", "-s"])
